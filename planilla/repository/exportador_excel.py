import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

from planilla.models import Empresa, Empleado
from planilla.constantes.tasas_css import calcular_css


class ExportadorExcel:
    """Genera el archivo .xlsx con hojas de Empresa y Empleados."""

    # ── Colores ───────────────────────────────────────────────────────────
    COLOR_PRIMARIO  = "1F4E79"
    COLOR_FILA_PAR  = "EBF5FB"
    COLOR_FILA_IMP  = "FDFEFE"
    COLOR_LABEL_BG  = "D6E4F0"

    def exportar(self, ruta: str, empresa: Empresa,
                 empleados: list[Empleado]) -> None:
        wb = openpyxl.Workbook()
        self._hoja_empresa(wb, empresa)
        self._hoja_empleados(wb, empleados)
        # eliminar hoja por defecto vacía
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(ruta)

    # ── Hoja Empresa ──────────────────────────────────────────────────────

    def _hoja_empresa(self, wb, empresa: Empresa) -> None:
        ws = wb.create_sheet("Empresa")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 38

        ws.merge_cells("A1:B1")
        t = ws["A1"]
        t.value = "📋  DATOS DE LA EMPRESA"
        t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        t.fill = PatternFill("solid", start_color=self.COLOR_PRIMARIO)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        campos = [
            ("Nombre",              empresa.nombre),
            ("RUC",                 empresa.ruc),
            ("DV",                  empresa.dv),
            ("Dirección",           empresa.direccion),
            ("Teléfono 1",          empresa.telefono1),
            ("Teléfono 2",          empresa.telefono2),
            ("Email",               empresa.email),
            ("Correlativo",         empresa.correlativo),
            ("Fecha de Apertura",   empresa.fecha_apertura),
            ("Fecha de Expiración", empresa.fecha_expiracion),
        ]

        for fila, (label, valor) in enumerate(campos, start=2):
            ws.row_dimensions[fila].height = 22
            lc = ws.cell(row=fila, column=1, value=label)
            vc = ws.cell(row=fila, column=2, value=valor)
            self._estilo_label(lc)
            self._estilo_valor(vc)

    # ── Hoja Empleados ────────────────────────────────────────────────────

    def _hoja_empleados(self, wb, empleados: list[Empleado]) -> None:
        ws = wb.create_sheet("Empleados")

        NOMBRES_MES = ["1er Mes", "2do Mes", "3er Mes", "4to Mes", "5to Mes",
                       "6to Mes", "7mo Mes", "8vo Mes", "9no Mes", "10mo Mes",
                       "11mo Mes", "12mo Mes", "13er Mes"]

        encabezados = [
            "Empresa", "Código", "Nombre", "Apellido", "Cédula", "F. Nacimiento",
            "Sexo", "Estado Civil", "Teléfono 1", "Teléfono 2",
            "Dirección", "Email", "F. Inicio", "F. Terminación",
            "Tipo Cuenta", "Cuenta Banco", "Departamento",
            "Tipo Planilla", "Status", "Seguro Social",
            *NOMBRES_MES,
            "Salario Neto", "Salario Anual", "ISR", "SSE", "SSP", "SEE", "SEP",
        ]
        anchos = [20, 12, 18, 18, 14, 14, 10, 14, 14, 14,
                  28, 26, 13, 15, 14, 18, 18, 16, 12, 14,
                  *[12]*13,
                  14, 14, 12, 12, 12, 12, 12]

        # Título
        last_col = get_column_letter(len(encabezados))
        ws.merge_cells(f"A1:{last_col}1")
        t = ws["A1"]
        t.value = "👥  PLANILLA DE EMPLEADOS"
        t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        t.fill = PatternFill("solid", start_color=self.COLOR_PRIMARIO)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        for col, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
            c = ws.cell(row=2, column=col, value=enc)
            self._estilo_encabezado(c)
            ws.column_dimensions[get_column_letter(col)].width = ancho
        ws.row_dimensions[2].height = 24

        fill_par = PatternFill("solid", start_color=self.COLOR_FILA_PAR)
        fill_imp = PatternFill("solid", start_color=self.COLOR_FILA_IMP)
        borde = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        hoy = date.today()

        for fila, emp in enumerate(empleados, start=3):
            ws.row_dimensions[fila].height = 20
            fill = fill_par if fila % 2 == 0 else fill_imp

            # Calcular 13 meses acumulativos
            acumulado = 0.0
            vals_mes = []
            for m in range(1, 14):
                horas_mes = getattr(emp, f"horas_mes_{m}", "") or ""
                if horas_mes:
                    ganado = self._calcular_ganado_mes(
                        emp.salario, emp.horas_quincena,
                        horas_mes, emp.tipo_planilla)
                else:
                    try:
                        ganado = float(emp.salario)
                    except (ValueError, TypeError):
                        ganado = 0.0
                acumulado = round(acumulado + ganado, 2)
                vals_mes.append(acumulado)

            salario_neto = acumulado  # acumulado mes 13

            # Salario Anual (salario mensual × 13)
            try:
                salario_f = float(emp.salario)
                descontado_f = float(emp.descontado) if getattr(emp, "descontado", "") else 0.0
                salario_anual_f = round((salario_f - descontado_f) * 13, 2)
                salario_anual = salario_anual_f
            except (ValueError, TypeError):
                salario_anual_f = 0.0
                salario_anual = ""

            # ISR = (Salario Anual - 11000) × tasa
            try:
                base_isr = salario_anual_f - 11000
                if base_isr <= 0:
                    isr = 0.0
                elif base_isr <= 50000:
                    isr = round(base_isr * 0.15, 2)
                else:
                    isr = round(base_isr * 0.25, 2)
            except Exception:
                isr = ""

            # Seguros sobre salario mensual
            if emp.seguro_social:
                try:
                    base_ss = float(emp.salario) - (float(emp.descontado) if getattr(emp, "descontado", "") else 0.0)
                except (ValueError, TypeError):
                    base_ss = 0.0
                css = calcular_css(base_ss, hoy)
                sse = css["empleado"]
                ssp = css["patronal"]
                see = css["see"]
                sep = css["sep"]
            else:
                sse = ssp = see = sep = ""

            valores = [
                getattr(emp, "empresa", "") or "",
                emp.codigo, emp.nombre, emp.apellido, emp.cedula,
                emp.fecha_nacimiento, emp.sexo, emp.estado_civil,
                emp.telefono1, emp.telefono2, emp.direccion, emp.email,
                emp.fecha_inicio, emp.fecha_terminacion,
                emp.tipo_cuenta, emp.cuenta_banco,
                emp.departamento, emp.tipo_planilla, emp.status,
                "Sí" if emp.seguro_social else "No",
                *vals_mes,
                salario_neto, salario_anual, isr, sse, ssp, see, sep,
            ]
            for col, val in enumerate(valores, start=1):
                c = ws.cell(row=fila, column=col, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = fill
                c.border = borde
                c.alignment = Alignment(vertical="center")
                if isinstance(val, float):
                    c.number_format = '"$"#,##0.00'

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = (
            f"A2:{last_col}{max(3, len(empleados) + 2)}"
        )

    # ── Helpers de cálculo ────────────────────────────────────────────────

    def _calcular_ganado_mes(self, salario: str, horas_contrato: str,
                              horas_reales: str, tipo_planilla: str) -> float:
        """Ganado en un mes = salario - descuento por horas no trabajadas."""
        try:
            s  = float(salario)
            hc = int(horas_contrato)
            hr = int(horas_reales)
            if hc <= 0 or hr <= 0:
                return s

            def _rata(h: int) -> float:
                if h < 88:
                    return s / h
                elif h <= 104:
                    return s / (h * 2)
                else:
                    return s / h

            descuento = (hc - hr) * _rata(hc)
            return round(s - descuento, 2)
        except (ValueError, ZeroDivisionError):
            return 0.0

    # ── Helpers de estilo ─────────────────────────────────────────────────

    def _estilo_encabezado(self, celda, color: str = None) -> None:
        color = color or self.COLOR_PRIMARIO
        celda.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", start_color=color)
        celda.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        celda.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

    def _estilo_label(self, celda) -> None:
        celda.font = Font(name="Arial", bold=True, size=10,
                          color=self.COLOR_PRIMARIO)
        celda.fill = PatternFill("solid", start_color=self.COLOR_LABEL_BG)
        celda.alignment = Alignment(vertical="center")

    def _estilo_valor(self, celda) -> None:
        celda.font = Font(name="Arial", size=10)
        celda.alignment = Alignment(vertical="center")
        celda.border = Border(bottom=Side(style="thin", color="AAAAAA"))
